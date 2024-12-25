import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from PIL import Image, ImageTk
import os
import openpyxl
from datetime import datetime
# Define the path for the logo and data file (Excel)
LOGO_PATH = "amar.png"  # Replace with the path to your logo file
DATA_FILE = "crm_data.xlsx"  # Excel file to save and load CRM data
LOG_FILE = "changes_log.txt"  # Log file to store changes


# Function to load data from the spreadsheet
def load_data():
    if os.path.exists(DATA_FILE):
        wb = openpyxl.load_workbook(DATA_FILE)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            data.append(row)
        return data
    return []


# Function to save data to the spreadsheet
def save_data(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Leads"

    # Define headers
    headers = [
        "Company Name", "Address", "Website", "Contact Person Name", "Designation",
        "Mobile Number", "Landline Number", "Company GST", "Additional Contact Person 1 Name", 
        "Mobile", "Email ID", "Branch", "Additional Contact Person 2 Name", 
        "Mobile", "Email ID", "Branch", "Follow-Up Date", "Response", "Next Follow-Up",
        "Meeting Reminder", "Meeting Outcome", "Quotation Given", "Negotiation", "Closed", "PO Number", "PO Date", 
        "PO Value", "Products Details", "Delivery Date"
    ]
    sheet.append(headers)

    # Append lead data to the sheet
    for lead in data:
        sheet.append(lead)

    wb.save(DATA_FILE)


# Function to log changes
def log_change(change_description):
    with open(LOG_FILE, "a") as log_file:
        log_file.write(f"{datetime.now()} - {change_description}\n")


# Main CRM Class
class SalesCRM:
    def __init__(self, root):
        self.root = root
        self.root.title("Sales CRM")
        self.data = load_data()

        # Display the logo at the top
        self.display_logo()

        # Frame for actions
        action_frame = tk.Frame(self.root)
        action_frame.pack(pady=10)

        tk.Button(action_frame, text="Add Lead", command=self.add_lead_popup).grid(row=0, column=0, padx=5)
        tk.Button(action_frame, text="View Leads", command=self.view_leads).grid(row=0, column=1, padx=5)
        tk.Button(action_frame, text="Update Follow-Up", command=self.update_followup_popup).grid(row=0, column=2, padx=5)
        tk.Button(action_frame, text="Add PO", command=self.add_po_popup).grid(row=0, column=3, padx=5)
        tk.Button(action_frame, text="Log View", command=self.view_log).grid(row=0, column=4, padx=5)

        # Frame for lead display
        self.display_frame = tk.Frame(self.root)
        self.display_frame.pack()

    # Method to display the logo
    def display_logo(self):
        try:
            image = Image.open(LOGO_PATH)
            image = image.resize((300, 100))  # Adjust dimensions as needed
            logo = ImageTk.PhotoImage(image)
            logo_label = tk.Label(self.root, image=logo)
            logo_label.image = logo  # Keep reference
            logo_label.pack(pady=10)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load logo: {e}")

    # Method to add lead
    def add_lead_popup(self):
        self.add_lead_window = tk.Toplevel(self.root)
        self.add_lead_window.title("Add Lead")
        
        # Entry fields for lead details
        labels = [
            "Company Name", "Address", "Website", "Contact Person Name", "Designation",
            "Mobile Number", "Landline Number", "Company GST", "Additional Contact Person 1 Name", 
            "Mobile", "Email ID", "Branch", "Additional Contact Person 2 Name", 
            "Mobile", "Email ID", "Branch", "Follow-Up Date", "Response", "Next Follow-Up",
            "Meeting Reminder", "Meeting Outcome", "Quotation Given", "Negotiation", "Closed", "PO Number", "PO Date", 
            "PO Value", "Products Details", "Delivery Date"
        ]
        
        self.entries = {}
        for i, label in enumerate(labels):
            tk.Label(self.add_lead_window, text=label).grid(row=i, column=0)
            entry = tk.Entry(self.add_lead_window, width=40)
            entry.grid(row=i, column=1)
            self.entries[label] = entry
        
        # Button to save lead
        tk.Button(self.add_lead_window, text="Save Lead", command=self.save_lead).grid(row=len(labels), columnspan=2, pady=10)

    # Method to save the lead to data
    def save_lead(self):
        lead = []
        for label in self.entries:
            lead.append(self.entries[label].get())

        self.data.append(lead)  # Add lead data to internal list
        save_data(self.data)  # Save to Excel file
        log_change("Added new lead")  # Log the change
        self.add_lead_window.destroy()  # Close the Add Lead popup
        messagebox.showinfo("Success", "Lead added successfully!")

    # Method to view leads
    def view_leads(self):
        view_window = tk.Toplevel(self.root)
        view_window.title("View Leads")

        columns = [
            "Lead Index", "Company Name", "Address", "Website", "Contact Person Name", "Designation",
            "Mobile Number", "Landline Number", "Company GST", "Follow-Up Date", "Response", 
            "Next Follow-Up", "Meeting Reminder", "PO Number"
        ]

        tree = ttk.Treeview(view_window, columns=columns, show="headings")
        tree.grid(row=0, column=0, padx=5, pady=5)

        # Define headings
        for col in columns:
            tree.heading(col, text=col)

        # Add leads to treeview
        for index, lead in enumerate(self.data, start=1):
            lead_with_index = [str(index)] + lead[:len(columns)-1]  # Add lead index at the beginning
            tree.insert("", "end", values=lead_with_index)

    # Method to view the change log
    def view_log(self):
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, "r") as log_file:
                log_content = log_file.readlines()
        else:
            log_content = ["No logs available."]
        
        log_window = tk.Toplevel(self.root)
        log_window.title("View Change Log")

        log_text = tk.Text(log_window, width=80, height=20)
        log_text.pack(padx=10, pady=10)
        log_text.insert(tk.END, "".join(log_content))
        log_text.config(state=tk.DISABLED)  # Make the text box read-only

    # Method to update follow-up
    def update_followup_popup(self):
        self.update_followup_window = tk.Toplevel(self.root)
        self.update_followup_window.title("Update Follow-Up")

        tk.Label(self.update_followup_window, text="Enter Lead Index to Update Follow-Up:").grid(row=0, column=0, padx=5, pady=5)
        lead_index_entry = tk.Entry(self.update_followup_window)
        lead_index_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(self.update_followup_window, text="Enter New Follow-Up Date:").grid(row=1, column=0, padx=5, pady=5)
        follow_up_date_entry = tk.Entry(self.update_followup_window)
        follow_up_date_entry.grid(row=1, column=1, padx=5, pady=5)

        def update_followup():
            try:
                lead_index = int(lead_index_entry.get()) - 1  # Convert to 0-based index
                new_follow_up_date = follow_up_date_entry.get()

                if lead_index < 0 or lead_index >= len(self.data):
                    raise IndexError("Invalid Lead Index.")

                # Update the follow-up date
                self.data[lead_index][17] = new_follow_up_date  # Follow-Up Date is at index 17
                save_data(self.data)
                log_change(f"Updated follow-up for Lead Index {lead_index + 1}")
                messagebox.showinfo("Success", f"Follow-up for Lead Index {lead_index + 1} updated successfully.")
                self.update_followup_window.destroy()

            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid Lead Index and Follow-Up Date.")
            except IndexError as e:
                messagebox.showerror("Invalid Lead Index", str(e))

        tk.Button(self.update_followup_window, text="Update Follow-Up", command=update_followup).grid(row=2, columnspan=2, pady=10)

    # Method to add purchase order (PO)
    def add_po_popup(self):
        self.add_po_window = tk.Toplevel(self.root)
        self.add_po_window.title("Add PO")

        tk.Label(self.add_po_window, text="Enter Lead Index to Add PO:").grid(row=0, column=0, padx=5, pady=5)
        lead_index_entry = tk.Entry(self.add_po_window)
        lead_index_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(self.add_po_window, text="Enter PO Number:").grid(row=1, column=0, padx=5, pady=5)
        po_number_entry = tk.Entry(self.add_po_window)
        po_number_entry.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(self.add_po_window, text="Enter PO Date:").grid(row=2, column=0, padx=5, pady=5)
        po_date_entry = tk.Entry(self.add_po_window)
        po_date_entry.grid(row=2, column=1, padx=5, pady=5)

        tk.Label(self.add_po_window, text="Enter PO Value:").grid(row=3, column=0, padx=5, pady=5)
        po_value_entry = tk.Entry(self.add_po_window)
        po_value_entry.grid(row=3, column=1, padx=5, pady=5)

        tk.Label(self.add_po_window, text="Enter Products Details:").grid(row=4, column=0, padx=5, pady=5)
        products_details_entry = tk.Entry(self.add_po_window)
        products_details_entry.grid(row=4, column=1, padx=5, pady=5)

        tk.Label(self.add_po_window, text="Enter Delivery Date:").grid(row=5, column=0, padx=5, pady=5)
        delivery_date_entry = tk.Entry(self.add_po_window)
        delivery_date_entry.grid(row=5, column=1, padx=5, pady=5)

        def add_po():
            try:
                lead_index = int(lead_index_entry.get()) - 1  # Convert to 0-based index
                po_number = po_number_entry.get()
                po_date = po_date_entry.get()
                po_value = po_value_entry.get()
                products_details = products_details_entry.get()
                delivery_date = delivery_date_entry.get()

                if lead_index < 0 or lead_index >= len(self.data):
                    raise IndexError("Invalid Lead Index.")
                
                # Update the PO details for the lead
                self.data[lead_index][23] = po_number  # PO Number
                self.data[lead_index][24] = po_date  # PO Date
                self.data[lead_index][25] = po_value  # PO Value
                self.data[lead_index][26] = products_details  # Products Details
                self.data[lead_index][27] = delivery_date  # Delivery Date

                save_data(self.data)
                log_change(f"Added PO for Lead Index {lead_index + 1}")
                messagebox.showinfo("Success", f"PO added for Lead Index {lead_index + 1}.")
                self.add_po_window.destroy()

            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid data for all fields.")
            except IndexError as e:
                messagebox.showerror("Invalid Lead Index", str(e))

        tk.Button(self.add_po_window, text="Add PO", command=add_po).grid(row=6, columnspan=2, pady=10)


# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = SalesCRM(root)
    root.mainloop()

